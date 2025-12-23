from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
import os
import tempfile
import re
from difflib import SequenceMatcher
import logging
from datetime import datetime

app = Flask(__name__)
CORS(app)

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 全局变量存储处理结果
process_result = {}

class ExcelSorter:
    def __init__(self):
        self.modules = []
        self.requirements = []
        self.unmatched_items = []
        self.processed_data = None

    def load_excel(self, file_path):
        """加载Excel文件"""
        try:
            # 检查文件是否存在
            if not os.path.exists(file_path):
                raise Exception(f"文件不存在: {file_path}")

            # 检查文件大小
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                raise Exception("文件为空")

            logger.info(f"开始读取Excel文件: {file_path} (大小: {file_size} bytes)")

            # 尝试不同的引擎读取Excel文件
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
            except Exception as e1:
                logger.warning(f"使用openpyxl引擎失败，尝试xlrd引擎: {str(e1)}")
                try:
                    df = pd.read_excel(file_path, engine='xlrd')
                except Exception as e2:
                    logger.error(f"使用xlrd引擎也失败: {str(e2)}")
                    raise Exception(f"无法读取Excel文件，尝试了openpyxl和xlrd引擎: {str(e1)}")

            # 验证数据
            if df is None or df.empty:
                raise Exception("Excel文件为空或无法读取数据")

            logger.info(f"成功加载Excel文件，共{len(df)}行{len(df.columns)}列数据")
            return df

        except Exception as e:
            logger.error(f"加载Excel文件失败: {str(e)}")
            raise Exception(f"无法读取Excel文件: {str(e)}")

    def extract_modules_and_requirements(self, df):
        """提取二级模块和需求名称（精确匹配版本）"""
        try:
            # A列是二级模块，D列是需求名称
            # 获取所有非空值，不去重，保留重复项
            modules = []
            if len(df.columns) > 0:
                for val in df.iloc[:, 0]:
                    if pd.notna(val) and str(val).strip() and str(val).strip() != 'nan':
                        modules.append(str(val).strip())

            # 获取需求名称 - 需要去重，确保每个需求只能被匹配一次
            requirements_set = set()
            requirements = []
            if len(df.columns) > 3:
                for val in df.iloc[:, 3]:
                    if pd.notna(val) and str(val).strip() and str(val).strip() != 'nan':
                        clean_req = str(val).strip()
                        # 只有未出现过的需求才添加到列表
                        if clean_req not in requirements_set:
                            requirements_set.add(clean_req)
                            requirements.append(clean_req)

            # 模块不去重，保留所有出现顺序；需求去重，避免重复匹配
            self.modules = modules
            self.requirements = requirements

            logger.info(f"提取到{len(self.modules)}个模块（包含重复项），{len(self.requirements)}个需求（去重）")
            return True
        except Exception as e:
            logger.error(f"提取模块和需求失败: {str(e)}")
            raise Exception(f"数据提取失败: {str(e)}")

    def calculate_similarity(self, str1, str2):
        """计算两个字符串的相似度"""
        return SequenceMatcher(None, str1.lower(), str2.lower()).ratio()

    def find_exact_match(self, module_value, requirements_list):
        """为二级模块值在需求列表中找到完全相同的匹配项"""

        # 去除首尾空格进行精确匹配
        cleaned_module = str(module_value).strip()

        for i, requirement in enumerate(requirements_list):
            cleaned_requirement = str(requirement).strip()
            # 精确匹配：文本必须完全相同
            if cleaned_module == cleaned_requirement:
                return i, cleaned_requirement, 'exact_match'

        # 没有找到完全匹配的项
        return None, None, 'no_match'

    def extract_keywords(self, text):
        """提取文本中的关键词"""
        # 移除常见的功能词汇
        stop_words = {'功能', '需求', '管理', '模块', '系统', '设置', '实现', '方案', '设计', '开发', '处理', '流程', '集成'}
        # 提取中文词汇（简单实现）
        keywords = []
        for char in text:
            if '\u4e00' <= char <= '\u9fff':  # 中文字符范围
                keywords.append(char)

        # 过滤停用词
        return [kw for kw in keywords if kw not in stop_words]

    def sort_requirements(self):
        """实现精确匹配：以二级模块为基准，在需求名称中查找完全相同的文本"""
        matched_results = []
        unmatched_modules = []
        matched_requirements_list = []

        # 复制需求列表，用于跟踪已匹配的项
        available_requirements = self.requirements.copy()

        # 为每个二级模块值查找精确匹配
        for module_value in self.modules:
            cleaned_module = str(module_value).strip()

            # 在可用需求中查找精确匹配
            match_index, matched_requirement, match_type = self.find_exact_match(
                cleaned_module, available_requirements
            )

            if match_type == 'exact_match' and match_index is not None:
                # 找到精确匹配
                matched_results.append({
                    'module_value': cleaned_module,
                    'matched_requirement': matched_requirement,
                    'match_status': '✅已匹配',
                    'match_method': '精确匹配'
                })

                # 记录已匹配的需求
                matched_requirements_list.append(matched_requirement)

                # 从可用列表中移除已匹配的项
                available_requirements.pop(match_index)

            else:
                # 没有找到匹配 - 保持空值，不填充任何内容
                matched_results.append({
                    'module_value': cleaned_module,
                    'matched_requirement': '',  # 保持为空
                    'match_status': '⚠️未匹配',
                    'match_method': '无'
                })

                # 记录未匹配的模块
                unmatched_modules.append(cleaned_module)

        # 获取所有未匹配的需求（在需求列表中但未被匹配的项）
        original_requirements = set(self.requirements)
        matched_requirements_set = set(matched_requirements_list)
        unmatched_requirements = list(original_requirements - matched_requirements_set)

        # 统计信息
        matched_count = len([r for r in matched_results if r['match_status'] == '✅已匹配'])
        total_count = len(matched_results)

        # 保存详细的匹配和未匹配信息
        self.matched_results = matched_results
        self.unmatched_modules = unmatched_modules
        self.unmatched_requirements = unmatched_requirements

        # 保存未匹配项信息（保持向后兼容）
        self.unmatched_items = [
            {
                'module_value': module,
                'reason': '在需求名称列中未找到完全相同的文本'
            }
            for module in unmatched_modules
        ]

        logger.info(f"精确匹配完成：成功匹配{matched_count}/{total_count}个二级模块")
        logger.info(f"未匹配模块：{len(unmatched_modules)}个，未匹配需求：{len(unmatched_requirements)}个")
        return matched_results

    def create_result_excel(self, df, sorted_requirements, output_path):
        """创建精确匹配的结果Excel文件，只保留关键列"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font

            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active

            # 设置表头
            headers = ['二级模块', '匹配的需求名称', '匹配状态', '匹配方式', '未匹配内容']
            ws.append(headers)

            # 设置表头样式
            header_font = Font(bold=True)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font

            # 定义样式
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            orange_fill = PatternFill(start_color='FFE5B4', end_color='FFE5B4', fill_type='solid')
            red_font = Font(color='FF0000')

            # 填充数据
            highlight_rows = []  # 记录需要标黄的行

            # 设置未匹配内容
            if hasattr(self, 'unmatched_requirements') and self.unmatched_requirements:
                # 第一行数据行显示"未匹配内容"
                ws.cell(row=2, column=5, value="未匹配内容")

                # 从第二行开始，每个未匹配需求占一行
                for i, req in enumerate(self.unmatched_requirements, 0):
                    ws.cell(row=3 + i, column=5, value=req)
            else:
                # 如果没有未匹配需求，显示"未匹配内容（无）"
                ws.cell(row=2, column=5, value="未匹配内容（无）")

            # 填充数据行
            for row_idx, result in enumerate(sorted_requirements, 2):  # 从第2行开始
                module_value = result['module_value']
                matched_requirement = result['matched_requirement']
                match_status = result['match_status']
                match_method = result['match_method']

                # 填充前4列数据
                row_data = [module_value, matched_requirement, match_status, match_method]
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

                # 如果未匹配，记录需要标黄的行
                if match_status == '⚠️未匹配':
                    highlight_rows.append(row_idx)

            # 标记未匹配的行
            for row_idx in highlight_rows:
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # 对未匹配行标黄
                    if col_idx == 2:  # 匹配的需求名称列
                        cell.fill = orange_fill  # 使用浅橙色
                    if col_idx == 3:  # 匹配状态列
                        cell.fill = yellow_fill
                        cell.font = red_font

            # 调整列宽
            column_widths = [15, 25, 12, 12, 40]  # 每列的宽度
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width

            # 保存文件
            wb.save(output_path)

            logger.info(f"精确匹配结果文件已保存至: {output_path}")
            return True

        except Exception as e:
            logger.error(f"保存结果文件失败: {str(e)}")
            raise Exception(f"保存结果失败: {str(e)}")

@app.route('/api/sort-excel', methods=['POST'])
def sort_excel():
    """处理Excel排序请求"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '不支持的文件格式'}), 400

        # 验证文件大小
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)

        if file_size == 0:
            return jsonify({'error': '文件为空'}), 400

        if file_size > 50 * 1024 * 1024:  # 50MB
            return jsonify({'error': '文件大小超过50MB限制'}), 400

        logger.info(f"开始处理文件: {file.filename}, 大小: {file_size} bytes")

        # 创建临时文件
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_input:
                file.save(tmp_input.name)
                input_path = tmp_input.name
                logger.info(f"临时文件已保存: {input_path}")
        except Exception as e:
            logger.error(f"保存临时文件失败: {str(e)}")
            return jsonify({'error': f'文件保存失败: {str(e)}'}), 500

        # 创建输出文件路径
        output_path = tempfile.mktemp(suffix='.xlsx')
        logger.info(f"输出文件路径: {output_path}")

        try:
            # 处理Excel文件
            sorter = ExcelSorter()

            # 加载数据
            logger.info("开始加载Excel文件...")
            df = sorter.load_excel(input_path)

            if df.empty:
                return jsonify({'error': 'Excel文件为空或格式不正确'}), 400

            logger.info(f"Excel文件加载成功，行数: {len(df)}, 列数: {len(df.columns)}")

            # 验证必要的列是否存在
            if len(df.columns) < 4:
                return jsonify({'error': 'Excel文件格式不正确：至少需要4列数据（A列：二级模块，D列：需求名称）'}), 400

            sorter.extract_modules_and_requirements(df)

            if not sorter.modules:
                return jsonify({'error': '未找到有效的二级模块数据（A列）'}), 400

            if not sorter.requirements:
                return jsonify({'error': '未找到有效的需求数据（D列）'}), 400

            # 排序需求
            logger.info("开始排序需求...")
            sorted_requirements = sorter.sort_requirements()

            if not sorted_requirements:
                return jsonify({'error': '需求排序失败'}), 500

            # 创建结果文件
            logger.info("生成结果文件...")
            sorter.create_result_excel(df, sorted_requirements, output_path)

            # 验证输出文件是否生成成功
            if not os.path.exists(output_path):
                return jsonify({'error': '结果文件生成失败'}), 500

            # 保存处理结果
            global process_result
            total_requirements = len(sorter.requirements)
            matched_requirements = total_requirements - len(sorter.unmatched_items)
            match_rate = round((matched_requirements / total_requirements * 100) if total_requirements > 0 else 0, 2)

            process_result = {
                'totalRequirements': total_requirements,
                'matchedRequirements': matched_requirements,
                'unmatchedRequirements': len(sorter.unmatched_items),
                'matchRate': match_rate,
                'details': [
                    {
                        'status': 'success',
                        'message': '成功匹配的需求',
                        'count': matched_requirements
                    },
                    {
                        'status': 'warning',
                        'message': '未能匹配的模块',
                        'count': len(sorter.unmatched_items)
                    },
                    {
                        'status': 'info',
                        'message': '处理的模块数量',
                        'count': len(sorter.modules)
                    }
                ],
                'warnings': [item['reason'] for item in sorter.unmatched_items] if sorter.unmatched_items else [],
                'unmatchedModules': getattr(sorter, 'unmatched_modules', []),
                'unmatchedRequirements': getattr(sorter, 'unmatched_requirements', []),
                'matchedDetails': getattr(sorter, 'matched_results', [])
            }

            logger.info(f"处理完成：总需求{total_requirements}个，成功匹配{matched_requirements}个")

            # 返回结果文件
            return send_file(
                output_path,
                as_attachment=True,
                download_name=f'sorted_{file.filename}',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except Exception as e:
            logger.error(f"处理Excel文件时出错: {str(e)}")
            import traceback
            traceback.print_exc()

            # 清理临时文件
            try:
                if os.path.exists(output_path):
                    os.unlink(output_path)
            except:
                pass

            return jsonify({'error': f'处理失败: {str(e)}'}), 500

        finally:
            # 清理输入临时文件
            try:
                if os.path.exists(input_path):
                    os.unlink(input_path)
                    logger.info("输入临时文件已清理")
            except Exception as e:
                logger.warning(f"清理输入临时文件失败: {str(e)}")

    except Exception as e:
        logger.error(f"处理请求时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'请求处理失败: {str(e)}'}), 500

@app.route('/api/process-result', methods=['GET'])
def get_process_result():
    """获取处理结果"""
    global process_result
    if process_result:
        return jsonify(process_result)
    else:
        return jsonify({'error': '没有处理结果'}), 404

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)